import tkinter as tk
from tkinter import messagebox
from threading import Thread
import time  
import os
import sys
import traceback
import getpass
from datetime import datetime
from playwright.sync_api import sync_playwright
import shutil
import pandas as pd
from openpyxl import load_workbook
import xlwings as xw
from openpyxl.utils import get_column_letter
import logging
from logging.handlers import RotatingFileHandler

# === GLOBAL CONFIG === 

date_columns = [
    "Service Date",	
    "Transaction Date",
    "Due Date",
    "Creation Date",
    "Earliest Due Date",
    "transaction_date",
    "service_date",
    "due_date",
    "creation date"
]
# Read Input.txt ONCE at the top ===
with open("Input.txt", "r") as f:
    lines = f.read().splitlines()
    username = lines[0]
    password = lines[1]
    save_dir_base = lines[2]

save_dir = os.path.join(save_dir_base, "Weekly Files")
date_str = datetime.today().strftime('%d-%b-%Y')
downloads_dir = save_dir #this points to the weekly path

# --- Logging setup ---
now = datetime.now()
job_name = now.strftime("Job_Run_%Y%m%d_%H%M%S")
log_timestamp = job_name
logs_dir = os.path.join(save_dir_base, "Logs")
os.makedirs(logs_dir, exist_ok=True)  # Ensure Logs directory exists
log_file = os.path.join(logs_dir, f"{log_timestamp}_log.txt")

# Chromium path setup
def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.abspath(relative_path)

chromium_path = resource_path("chrome-win/chrome.exe")

# Configure logging
logger = logging.getLogger("ar_aging")
logger.setLevel(logging.INFO)

file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
file_handler.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))

console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))

logger.addHandler(file_handler)
logger.addHandler(console_handler)
# === END OF GLOBAL CONFIG ===


# === Function for APP Run 1 - Run report ===
def getreport_job():
    try:
        logger.info("*************Starting Report run *****************")
        # def resource_path(relative_path):
        #     if hasattr(sys, "_MEIPASS"):
        #         return os.path.join(sys._MEIPASS, relative_path)
        #     return os.path.abspath(relative_path)

        # #chromium_path = resource_path("chrome-win/chrome.exe")

        def handle_dialog(dialog):
            logger.info(f"Popup: {dialog.message}")
            dialog.accept()

        def fill_Australia(page, retries=3):
            for attempt in range(retries):
                logger.info(f"🔁 Attempt {attempt + 1} to select Australia...")
                for frame in page.frames:
                    if frame.locator("text=Business Unit").is_visible(timeout=1000):
                        logger.info("Found 'Business Unit'")
                        frame.locator("text=Business Unit").click()
                    if frame.locator("text=AllAUSTRALIANEW ZEALANDSearch").is_visible(timeout=1000):
                        logger.info("Found dropdown options container")
                        frame.locator("text=AllAUSTRALIANEW ZEALANDSearch").click()
                    if frame.locator("label span:text('AUSTRALIA')").is_visible(timeout=1000):
                        logger.info("Found 'AUSTRALIA' checkbox")
                        checkbox = frame.locator("input[value='AUSTRALIA']")
                        if not checkbox.is_checked():
                            checkbox.check()
                        logger.info("AUSTRALIA selected.")
                        return True
                time.sleep(2)
            return False

        def try_to_click_notification_link(page, retries=3):
            for attempt in range(retries):
                logger.info(f"🔁 Attempt {attempt + 1} to click 'Notification' link...")
                for frame in page.frames:
                    try:
                        link = frame.locator("a:has-text('Notification')")
                        if link.count() > 0:
                            link.nth(0).click()
                            logger.info("✅ Clicked 'Notification'.")
                            return True
                    except Exception:
                        logger.exception("Exception while clicking notification link")
                time.sleep(2)
            return False

        def fill_email_checkbox_and_submit(page, retries=3):
            for attempt in range(retries):
                logger.info(f"🔁 Attempt {attempt + 1} to fill email and submit...")
                for frame in page.frames:
                    try:
                        if frame.locator("input[type='checkbox'][name='email_notification']").is_visible():
                            frame.locator("input[type='checkbox'][name='email_notification']").click()
                            logger.info("✅ Checkbox clicked.")
                            break
                    except Exception:
                        logger.exception("Exception while clicking checkbox")
                for frame in page.frames:
                    try:
                        if frame.locator("input[name='notify_to']").is_visible():
                            frame.locator("input[name='notify_to']").fill(username)
                            logger.info("✅ Email textbox filled.")
                            break
                    except Exception:
                        logger.exception("Exception while filling email textbox")
                for frame in page.frames:
                    try:
                        if frame.locator("button[name='submitButton']").is_visible():
                            frame.locator("button[name='submitButton']").click()
                            logger.info("✅ Submit button clicked.")
                            return True
                    except Exception:
                        logger.exception("Exception while clicking submit button")
                time.sleep(2)
            return False

        def fill_jobname_and_pressok(page, retries=3):
            for attempt in range(retries):
                logger.info(f"🔁 Attempt {attempt + 1} to fill job name and OK...")
                for frame in page.frames:
                    try:
                        if frame.locator("input[name='submitJobName']").is_visible():
                            frame.locator("input[name='submitJobName']").fill(job_name)
                            logger.info("✅ Job name filled.")
                            break
                    except Exception:
                        logger.exception("Exception while filling job name")
                for frame in page.frames:
                    try:
                        if frame.locator("#submitDiv_button").is_visible():
                            frame.locator("#submitDiv_button").click()
                            logger.info("✅ OK clicked.")
                            return True
                    except Exception:
                        logger.exception("Exception while clicking OK")
                time.sleep(2)
            return False

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, executable_path=chromium_path)
            context = browser.new_context()
            page = context.new_page()
            page.on("dialog", handle_dialog)

            logger.info("🌐 Navigating to report URL...")
            page.goto("https://efyq.fa.ap4.oraclecloud.com/analytics/saw.dll?bipublisherEntry&Action=schedule&itemType=.xdo&bipPath=%2FCustom%2FFinancials%2FReceivables%2FAR%20Aging.xdo&path=%2Fshared%2FCustom%2FFinancials%2FReceivables%2FAR%20Aging.xdo")
            logger.info("✅ Report page loaded.")

            page.get_by_role("button", name="Company Single Sign-On").click()
            page.get_by_role("textbox", name="FirstName.LastName@invocare.").fill(username)
            page.get_by_role("button", name="Next").click()
            page.get_by_role("textbox", name="Enter the password for").fill(password)
            page.get_by_role("button", name="Sign in").click()
            logger.info("✅ Credentials submitted.")

            page.wait_for_load_state("load")
            time.sleep(3)

            if not fill_Australia(page):
                raise Exception("Failed to select Australia.")
            if not try_to_click_notification_link(page):
                raise Exception("Failed to click 'Notification' link.")
            if not fill_email_checkbox_and_submit(page):
                raise Exception("Failed to fill checkbox and submit.")
            if not fill_jobname_and_pressok(page):
                raise Exception("Failed to fill job name and press OK.")

            browser.close()

    except Exception as e:
        logger.error("❌ Unhandled Exception:")
        logger.exception(e)
        raise

    logger.info(f"✅ Job completed successfully.\nLog saved to: {log_file}")
    return True, "Job completed successfully."





# --- Link clicking logic - Used in App run 2 - download report---
def try_to_click_first_link(page, retry_attempts=3, retry_delay=3):
    for attempt in range(1, retry_attempts + 1):
        logger.info(f"Attempt {attempt} to locate and click the 'first' link...")

        frames = page.frames
        if not frames:
            logger.warning("No frames found on the page.")
            return False

        for frame in frames:
            logger.info(f"Checking frame: {frame.name}")
            try:
                notification_links = frame.locator("span[class='tabBottomT2LMargin2']")
                links_count = notification_links.count()

                if links_count == 0:
                    logger.info(f"No 'first' links found in frame {frame.name}.")
                    continue

                notification_links.nth(0).click()
                logger.info(f"Clicked 'first' link in frame {frame.name}.")
                return True
            except Exception as e:
                logger.error(f"Error while checking frame {frame.name}: {e}")
                logger.exception(e)

        logger.info(f"Retrying in {retry_delay} seconds...")
        time.sleep(retry_delay)

    logger.warning(f"Could not click any 'first' link after {retry_attempts} attempts.")
    return False

# --- Downloading logic - Used in App run 2 - download report----
def try_to_download_first_link(page, retry_attempts=3, retry_delay=3):
    target_filename = datetime.now().strftime("%d-%b-%Y_%H%M%S") + ".xls"

    for attempt in range(1, retry_attempts + 1):
        logger.info(f"Attempt {attempt} to locate and download the 'Output1' link...")

        frames = page.frames
        if not frames:
            logger.warning("No frames found on the page.")
            return False

        for frame in frames:
            logger.info(f"Checking frame: {frame.name}")

            try:
                link = frame.locator("span.dataLink").filter(has_text="Output1")
                if link.count() == 0:
                    logger.info(f"No 'Output1' link found in frame {frame.name}.")
                    continue

                with page.expect_download() as download_info:
                    link.first.click()

                download = download_info.value
                download_path = download.path()
                if not download_path:
                    logger.warning("Download path not found.")
                    continue

                new_path = os.path.join(downloads_dir, target_filename)
                download.save_as(new_path)
                logger.info(f"Downloaded file saved as: {new_path}")
                return True

            except Exception as e:
                logger.error(f"Error while downloading in frame {frame.name}: {e}")
                logger.exception(e)

        logger.info(f"Retrying in {retry_delay} seconds...")
        time.sleep(retry_delay)

    logger.warning(f"Could not download the 'Output1' link after {retry_attempts} attempts.")
    return False

# Final xlsx worksheets creation - Used in App run 2 - download report---
def Process_xlsx(file_path):
    # Read Excel file with fixed header in row 0
    raw_data = pd.read_excel(file_path, sheet_name=0, header=0, parse_dates=date_columns)
    
    # Define filter conditions
    bu = [
        "AUSTRALIA", 
        "AUSTRALIA", 
        "AUSTRALIA", 
        "AUSTRALIA", 
        "AUSTRALIA",
        "AUSTRALIA", 
        "AUSTRALIA", 
        "AUSTRALIA", 
        "AUSTRALIA",
        "AUSTRALIA", 
        "AUSTRALIA",
        "AUSTRALIA",
        "AUSTRALIA", 
        "AUSTRALIA"
    ]

    steps = [
    "Other Receipts",
    "Unidentified Receipts",
    "NSW & QLD Cem Crem",
    "Pinegrove",
    "Forest Lawn",
    "Funeral Directors Life Art",
    "Funerals WA SA VIC TAS",
    "Funerals NSW & QLD",
    "Debt Collectors",
    "Legal",
    "Interco",
    "Fund",
    "Missed 1",
    "Missed 2"
    ]



    acc_no = [
    "!3RYB, 4VKK",
    "3RYB, 4VKK",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*"
    ]


    zone = [
    "*",
    "*",
    "NSW Cem Crem, QLD CemCrem",
    "NSW Cem Crem",
    "NSW Cem Crem",
    "*",
    "VIC-TAS Funerals, WASA Funerals",
    "QLD NNSW Funerals, NSW Funerals",
    "*",
    "*",
    "*",
    "*",
    "(blank)",
    "!(blank)"
    ]



    profile = [
    "*",
    "*",
    "Dunning - Statement, Dunning -No Statement, CEM CREM, DEFAULT, Installment Payer - Statement, No Dunning, PRENEED, PROBATE, Statement - No Dunning",
    "Dunning - Statement, Dunning -No Statement, CEM CREM, DEFAULT, Installment Payer - Statement, No Dunning, PRENEED, PROBATE, Statement - No Dunning",
    "Dunning - Statement, Dunning -No Statement, CEM CREM, DEFAULT, Installment Payer - Statement, No Dunning, PRENEED, PROBATE, Statement - No Dunning",
    "EXTERNAL FD",
    "Dunning - Statement, Dunning -No Statement, DEFAULT, FUNERAL, Installment Payer - Statement, No Dunning, PROBATE, Statement - No Dunning",
    "Dunning - Statement, Dunning -No Statement, DEFAULT, FUNERAL, Installment Payer - Statement, No Dunning, PROBATE, Statement - No Dunning",
    "External Debt - Installment, External Debt Collector, External Debt Collector - pend, Write Off",
    "LEGAL FIRM, Legal Firm - No Dunning, Legal Firm - No Statement, Legal Firm - Statement",
    "INTERNAL FD",
    "TRUST",
    "!INTERNAL FD, EXTERNAL FD, External Debt - Installment, External Debt Collector, External Debt Collector - pending, LEGAL FIRM, Legal Firm - No Dunning, Legal Firm - No Statement, Legal Firm - Statement",
    "(blank)"
    ]



    branch = [
    "*",
    "*",
    "!Pinegrove Memorial Park - Admin, Forest Lawn Memorial Park - Admin",
    "Pinegrove Memorial Park - Admin",
    "Forest Lawn Memorial Park - Admin",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*",
    "*"
    ]


    t_type = [
    "Receipt, Credit Memo",
    "Receipt, Credit Memo",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice",
    "Invoice"
    ]


    # Use date and time for unique processed file name
    processed_timestamp = datetime.now().strftime('%d-%b-%Y_%H%M%S')
    output_file = os.path.join(save_dir, f"{processed_timestamp}_Processed.xlsx")
    shutil.copy(file_path, output_file)

    app = xw.App(visible=False)
    wb = app.books.open(output_file)
    original_sheet = wb.sheets[0]
    original_sheet.name = "Raw Data"

    #trial code
    def apply_filter(df, column_name, filter_value):
        if filter_value == "*" or column_name not in df.columns:
            return df

        # Normalize the column (strip whitespace, convert NaNs to empty string for matching)
        column_series = df[column_name].astype(str).str.strip()
        is_exclude = filter_value.startswith("!")
        
        raw_values = filter_value[1:] if is_exclude else filter_value
        values = [v.strip() for v in raw_values.split(",")]

        # Handle blank logic
        match_blanks = "(blank)" in values
        values = [v for v in values if v != "(blank)"]

        if is_exclude:
            mask = ~column_series.isin(values)
            if match_blanks:
                # Also exclude blanks (NaN or empty)
                blank_mask = df[column_name].isna() | (column_series == "")
                mask = mask & ~blank_mask
        else:
            mask = column_series.isin(values)
            if match_blanks:
                # Also include blanks
                blank_mask = df[column_name].isna() | (column_series == "")
                mask = mask | blank_mask

        return df[mask].copy()


    
    # Apply filters and write to separate new sheets
    for i, step in enumerate(steps):
        df_filtered = raw_data.copy()
        
        #then use
        columns_to_filter = [
            ("Account Number", acc_no[i]),
            ("Business Unit", bu[i]),
            ("Zone", zone[i]),
            ("Profile Class", profile[i]),
            ("Transaction Type", t_type[i]),
            ("Branch Name", branch[i]),
        ]   

        for col_name, filter_val in columns_to_filter:
            df_filtered = apply_filter(df_filtered, col_name, filter_val)
            logger.info(f"Step '{step}' - Filtered rows: {len(df_filtered)}")

       
        # Delete existing sheet with same name if it exists
        if step in [s.name for s in wb.sheets]:
            wb.sheets[step].delete()

        # Create new sheet and paste filtered data
        sht = wb.sheets.add(step)
        sht.range("A1").options(index=False).value = df_filtered

        # Format headers bold, disable wrap text, autofit
        rng = sht.range("A1").expand("table")
        rng.api.WrapText = False
        sht.range("A1").expand("right").api.Font.Bold = True
        #sht.autofit()

        # Find the last used column in the first row
        #last_col = sht.range('1:1').end('left').column
        # Set the header for the next column
        #sht.cells(1, last_col + 1).value = "User Comments"
    
    # Save and close
    wb.save()
    logger.info(f"Processed file saved")
    wb.close()
    app.quit()

    print("AR Aging Processing Completed! Output saved to:", output_file)

# --- Main Automation - RUN APP 2 ---
def start_Downloadprocess():
    try:
        logger.info("\n\n\n******* Starting AR Aging Download Process ***********")
        logger.info("Starting automation script...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, executable_path=chromium_path)
            context = browser.new_context(accept_downloads=True)
            page1 = context.new_page()

            logger.info("Navigating to report page...")
            page1.goto("https://efyq.fa.ap4.oraclecloud.com/analytics/saw.dll?bipublisherEntry&Action=history&itemType=.xdo&path=/shared/Custom/Financials/Receivables&bipPath=/Custom/Financials/Receivables/AR%20Aging.xdo")

            logger.info("Clicking SSO button...")
            page1.get_by_role("button", name="Company Single Sign-On").click()

            logger.info("Filling login email...")
            page1.get_by_role("textbox", name="FirstName.LastName@invocare.").fill(username)

            logger.info("Clicking Next...")
            page1.get_by_role("button", name="Next").click()

            logger.info("Entering password...")
            page1.get_by_role("textbox", name="Enter the password for").fill(password)

            logger.info("Clicking Sign in...")
            page1.get_by_role("button", name="Sign in").click()

            logger.info("Waiting for page to load...")
            page1.wait_for_load_state("load")
            time.sleep(3)

            logger.info("Looking for notification link...")
            if try_to_click_first_link(page1):
                logger.info("Notification link clicked.")

                logger.info("Looking for report download...")
                if try_to_download_first_link(page1):
                    logger.info("Report downloaded successfully.")
                else:
                    logger.warning("Failed to download report.")
                    raise FileNotFoundError("The AR report could not be downloaded from the browser. Please check the portal and try again.")
            else:
                logger.warning("Failed to find notification link.")
                raise FileNotFoundError("Could not find the notification link to download the AR report. Please check the portal and try again.")

            browser.close()
            logger.info("Browser closed.")

        # File conversion
        # Find the latest .xls file in the downloads_dir (Weekly Files)
        xls_files = [f for f in os.listdir(downloads_dir) if f.lower().endswith(".xls")]
        if not xls_files:
            logger.warning("No .xls files found in the Weekly Files folder.")
            return

        latest_xls = max(xls_files, key=lambda f: os.path.getmtime(os.path.join(downloads_dir, f)))
        source_path = os.path.join(downloads_dir, latest_xls)
        converted_timestamp = datetime.now().strftime('%d-%b-%Y_%H%M%S')
        new_file_path = os.path.join(save_dir, f"{converted_timestamp}_Converted.xlsx")  # Save with date_time_converted.xlsx
        #new_file_path = os.path.join(save_dir, f"{date_str}.xlsx")  # Save new file in save_dir_base

        logger.info(f"Converting latest file {latest_xls} to .xlsx format...")
        app = xw.App(visible=False)
        wb = app.books.open(source_path)

        wb.sheets[0].range("1:3").delete()

        wb.save(new_file_path)
        wb.close()
        app.quit()
        logger.info(f"File converted and saved to: {new_file_path}")
        Process_xlsx(new_file_path)
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        logger.exception(e)
        raise  # Let run_app3 handle the popup

#*****END OF DOWNLOAD REPORT CODE*********


# === Core Processing Logic ===
def is_file_locked(file_path):
    """Check if the file is open by another application."""
    try:
        with open(file_path, 'r+'):
            return False
    except (IOError, PermissionError):
        return True

def update_fd_customers(excel_path,wb, prev_master_path):
    # Load workbook and worksheet
    
    #ws_source = wb['Funeral Directors Life Art']
    ws_target = wb['FD Customers']

    # Read the source sheet into a DataFrame
    df_source = pd.read_excel(excel_path, sheet_name='Funeral Directors Life Art', parse_dates=date_columns)

    # Ensure required columns are present
    required_columns = ['Customer', 'Account Number']
    if not all(col in df_source.columns for col in required_columns):
        raise ValueError("One or all required columns not found in source sheet.")

    # Drop duplicates based on both columns
    df_filtered = df_source[required_columns].drop_duplicates().reset_index(drop=True)

    # Drop duplicates based only on 'Account Number', keep required columns
    df_filtered = df_source.drop_duplicates(subset=['Account Number'])[required_columns].reset_index(drop=True)

    # Find the column letters of 'Customer' and 'Account Number' in target sheet
    header_row = [cell.value for cell in ws_target[1]]
    col_indices = {col: header_row.index(col) + 1 for col in required_columns if col in header_row}

    if len(col_indices) < 2:
        raise ValueError("Could not find both 'Customer' and 'Account Number' columns in target sheet.")

    # Clear old data in those columns (starting from row 2)
    max_row = ws_target.max_row
    for col_letter in col_indices.values():
        for row in range(2, max_row + 1):
            ws_target.cell(row=row, column=col_letter).value = None

    # Write new deduplicated data into the right columns
    for i, row in df_filtered.iterrows():
        for col_name, col_index in col_indices.items():
            ws_target.cell(row=i + 2, column=col_index, value=row[col_name])

    # --- Copy User Comments from previous master file ---
    #from openpyxl import load_workbook

    ws_target = wb['FD Customers']

    # Load previous master workbook and FD Customers sheet
    wb_prev = load_workbook(prev_master_path, read_only=True)
    ws_prev = wb_prev['FD Customers']

    # Find column indices for 'Customer' and 'User Comments'
    header = [cell.value for cell in ws_target[1]]
    try:
        customer_col_idx = header.index('Customer') + 1
        comments_col_idx = header.index('User Comments') + 1
    except ValueError:
        logger.info(f"Required columns not found in FD Customers sheet.")
        return

    # Build mapping from Customer to User Comments in previous master
    prev_comments = {}
    for row in ws_prev.iter_rows(min_row=2, values_only=True):
        customer = row[customer_col_idx - 1]
        comment = row[comments_col_idx - 1]
        if customer:
            prev_comments[customer] = comment

    # Clear and update User Comments in current master
    for row in ws_target.iter_rows(min_row=2):
        customer = row[customer_col_idx - 1].value
        # Clear current comment
        row[comments_col_idx - 1].value = None
        # Copy from previous if exists
        if customer in prev_comments:
            row[comments_col_idx - 1].value = prev_comments[customer]

    wb_prev.close()



def update_top50_comments(master_file, prev_master_file, src_range, dest_cell):
    logger.info(f"=== update_top50_comments (xlwings) called ===")
    logger.info(f"Master file: {master_file}")
    logger.info(f"Previous master file: {prev_master_file}")

    app = xw.App(visible=False)
    try:
        wb_master = app.books.open(master_file)
        wb_prev = app.books.open(prev_master_file)

        for sheet_name in ["Top 50 Receipts", "Top 50 Invoice"]:
            if sheet_name not in [s.name for s in wb_master.sheets] or sheet_name not in [s.name for s in wb_prev.sheets]:
                logger.warning(f"Sheet '{sheet_name}' not found in one of the workbooks.")
                continue

            sht_master = wb_master.sheets[sheet_name]
            sht_prev = wb_prev.sheets[sheet_name]

            # Force calculation in both workbooks
            sht_master.api.Calculate()
            sht_prev.api.Calculate()

            # Copy values from src_range to dest_cell in master
            src_values = sht_master.range(src_range).value
            sht_master.range(dest_cell).value = src_values
            logger.info(f"✅ Copied values from {src_range} to {dest_cell} in '{sheet_name}'.")

            # Find header row and columns
            header = sht_master.range("1:1").value
            customer_col_idx = header.index("Customer") + 1
            comments_col_idx = header.index("User Comments") + 1

            # Build mapping from previous file: Customer -> User Comment
            prev_data = sht_prev.range("I1").expand("table").value
            prev_header = prev_data[0]
            prev_customer_idx = prev_header.index("Customer")
            prev_comments_idx = prev_header.index("User Comments")
            prev_comments_map = {}
            for row in prev_data[1:]:
                customer = row[prev_customer_idx]
                comment = row[prev_comments_idx]
                if customer is not None and str(customer).strip() != "":
                    prev_comments_map[str(customer).strip()] = comment

            # Get the destination range for customers and comments
            dest_start_row = sht_master.range(dest_cell).row
            dest_start_col = sht_master.range(dest_cell).column
            num_rows = sht_master.range(src_range).rows.count

            logger.info(f"Processing rows {dest_start_row} to {dest_start_row + num_rows - 1} in '{sheet_name}' for customer/comment update.")
            updated_count = 0
            for i in range(num_rows):
                row_idx = dest_start_row + i
                customer = sht_master.cells(row_idx, customer_col_idx).value
                sht_master.cells(row_idx, comments_col_idx).value = None  # Clear
                customer_key = str(customer).strip() if customer else ""
                if customer_key in prev_comments_map and prev_comments_map[customer_key] is not None:
                    sht_master.cells(row_idx, comments_col_idx).value = prev_comments_map[customer_key]
                    updated_count += 1

            logger.info(f"✅ Updated {updated_count} user comments in '{sheet_name}'.")

        wb_master.save()
        logger.info(f"✅ Master workbook saved after updating Top 50 comments: {master_file}")
        wb_master.close()
        wb_prev.close()
        logger.info("Closed both master and previous workbooks.")

    except Exception as e:
        logger.error(f"❌ Error in update_top50_comments (xlwings): {e}")
        logger.exception(e)
        try:
            wb_master.close()
        except Exception:
            pass
        try:
            wb_prev.close()
        except Exception:
            pass
    finally:
        app.quit()


#====== Master File Update Script - APP RUN 3 Code ====
# This function is called by the GUI button to run the Master File Update script
def run_Master_file_script():
    try:
        logger.info("\n\n\n*********** Starting Master File Update Script **********")
        logger.info("Picking path from input file")
        folder_path = save_dir_base  # Use global config
        logger.info(f"Path is {folder_path}")

        master_file = os.path.join(folder_path, "AR Aging Master File.xlsx")
        weekly_folder = os.path.join(folder_path, "Weekly Files")
        archive_folder = os.path.join(folder_path, "Master File Archive")
        transaction_type_col = "Transaction Type"
        invoice_number_col = "Invoice Number"
        receipt_number_col = "Receipt Number"
        comments_col = "User Comments"
        status_col = "Status"
        logger.info("Variables set")

        os.makedirs(archive_folder, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archived_file = os.path.join(archive_folder, f"AR Aging Master File {timestamp}.xlsx")
        shutil.copy2(master_file, archived_file)
        logger.info(f"✅ Archived master file saved as: {archived_file}")

        # Check if the files are locked
        if is_file_locked(master_file):
            logger.error(f"❌ The master file '{master_file}' is currently open. Please close it and try again.")
            messagebox.showerror("Error", f"The master file '{master_file}' is open. Please close it and try again.")
            return

        weekly_files = [os.path.join(weekly_folder, f) for f in os.listdir(weekly_folder) if f.endswith("Processed.xlsx")]
        if not weekly_files:
            raise FileNotFoundError("❌ No weekly Excel files found in 'Weekly Files' folder.")
        latest_weekly_file = max(weekly_files, key=os.path.getmtime)
        logger.info(f"📄 Using latest weekly file: {latest_weekly_file}")

        # Check if the weekly file is locked
        if is_file_locked(latest_weekly_file):
            logger.error(f"❌ The weekly file '{latest_weekly_file}' is currently open. Please close it and try again.")
            messagebox.showerror("Error", f"The weekly file '{latest_weekly_file}' is open. Please close it and try again.")
            return

        logger.info("Loading master workbook...")
        wb_master = load_workbook(master_file)
        logger.info("Master workbook loaded.")

        logger.info("Loading weekly workbook...")
        sheet_names = pd.ExcelFile(latest_weekly_file).sheet_names
        logger.info("Weekly workbook loaded.")

        # Update today's date in the "Others" worksheet in cell D2
        if "Others" in wb_master.sheetnames:
            ws_others = wb_master["Others"]
            ws_others["D2"] = datetime.today().strftime("%d-%b-%Y")
            logger.info("✅ Updated run date.")
            ws_others["D4"] = datetime.now().strftime("%H:%M:%S")  # Add current time to D4
            logger.info("✅ Updated run time.")
            ws_others["D5"] = getpass.getuser()  # Add user ID to D5
            logger.info("✅ Updated user ID.")
        else:
            logger.warning("⚠️ 'Others' worksheet not found in the master file.")

        
        def get_unique_id_column(row):
            ttype = str(row.get(transaction_type_col, "")).strip().lower()
            if ttype == "invoice" or ttype == "credit memo":
                return row.get(invoice_number_col)
            elif ttype == "receipt":
                return row.get(receipt_number_col)
            return None
       
        for sheet in sheet_names:
            if sheet == "XDO_METADATA":
                logger.info(f"⏭️ Skipping special sheet '{sheet}'")
                continue

            if sheet not in wb_master.sheetnames:
                logger.info(f"⏭️ Skipping sheet '{sheet}' — not found in master file.")
                continue

            logger.info(f"🔄 Processing sheet: {sheet}")

            df_weekly = pd.read_excel(latest_weekly_file, sheet_name=sheet, parse_dates=date_columns)
            df_master_old = pd.read_excel(archived_file, sheet_name=sheet, parse_dates=date_columns)

            df_weekly = df_weekly[df_weekly[transaction_type_col].notna()]
            df_master_old = df_master_old[df_master_old[transaction_type_col].notna()]

            df_weekly["__UniqueID__"] = df_weekly.apply(get_unique_id_column, axis=1)
            df_master_old["__UniqueID__"] = df_master_old.apply(get_unique_id_column, axis=1)

            df_weekly = df_weekly[df_weekly["__UniqueID__"].notna()]
            df_master_old = df_master_old[df_master_old["__UniqueID__"].notna()]

            df_merged = df_weekly.copy()

            user_input_columns = ["User Comments", "Status"]
            existing_user_input_cols = [col for col in user_input_columns if col in df_master_old.columns]

            if existing_user_input_cols:
                df_user = df_master_old[["__UniqueID__"] + existing_user_input_cols].drop_duplicates()
                df_merged = pd.merge(df_weekly, df_user, on="__UniqueID__", how="left")
                logger.info(f"✅ {', '.join(existing_user_input_cols)} copied from old master for sheet '{sheet}'.")
            else:
                df_merged = df_weekly.copy()
                logger.warning(f"⚠️ No user columns ({', '.join(user_input_columns)}) found in old master sheet '{sheet}' — skipping copy.")

            df_merged.drop(columns=["__UniqueID__"], inplace=True)

            ws = wb_master[sheet]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if not cell.data_type == "f":
                        cell.value = None

            logger.info(f"✅ Cleared old data in sheet '{sheet}'.")

            user_input_columns = ["User Comments", "Status"]  # You can change this

            # Read headers from row 1 in the worksheet
            master_headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    master_headers[header.strip()] = col
            logger.info(f"✅ Master headers loaded for sheet '{sheet}'")
            
            # Go row-by-row from the DataFrame new code
            for row_idx, row in enumerate(df_merged.itertuples(index=False), start=2):
                for col_name, value in zip(df_merged.columns, row):
                    col_name = col_name.strip()
                    if col_name in master_headers:
                        col_idx = master_headers[col_name]
                        # If this is a date column and value is not null
                        if col_name in date_columns and pd.notnull(value):
                            if not isinstance(value, datetime):
                                value = pd.to_datetime(value)
                            if hasattr(value, "tzinfo") and value.tzinfo is not None:
                                value = value.replace(tzinfo=None)
                        ws.cell(row=row_idx, column=col_idx, value=value)
            logger.info(f"✅ Data written to sheet '{sheet}'.")

        # copy paste values for top 50 sheets
        # copy_range_openpyxl(
        #     file_path="your_file.xlsx",
        #     sheet_name="Sheet1",
        #     src_range="A1:C5",
        #     dest_cell="E1"
        #     )

        # Update FD Customers in master file      
        update_fd_customers(master_file, wb_master, archived_file)
        logger.info("✅ FD Customers updated in master file.")
        
        

        wb_master.save(master_file)
        wb_master.close()
        logger.info("✅ Master file updated with latest data and preserved user comments.")

        # Update user comments in Top 50 sheets
        update_top50_comments(master_file, archived_file,"D2:G51", "I2")
        logger.info("✅ User comments updated in Top 50 sheets.")

        logger.info(f"✅ Master file saved at: {master_file}")
        logger.info(f"✅ Log file for this run: {log_file}")
        logger.info("✅ App run complete.")

    except Exception as e:
        logger.error("❌ An error occurred during execution.")
        logger.exception(e)
        raise  # Let run_app3 handle the popup

#*****END OF UPDATE MASTER FILE CODE*********



#******Consolidate comments to Raw data APP RUN 4 NOT BEING USED*********

def consolidate_comments_to_raw_data():
    try:
        logger.info("Starting consolidation of User Comments and Status to Raw Data...")

        base_dir = save_dir_base
        master_file = os.path.join(base_dir, "AR Aging Master File.xlsx")
        logger.info(f"Opening master file: {master_file}")
        
        logger.info("Loading workbook...")
        wb = load_workbook(master_file)
        logger.info(f"Workbook loaded with {len(wb.sheetnames)} sheets")
        
        if "Raw Data" not in wb.sheetnames:
            logger.error("❌ Error: Raw Data sheet not found in the master file.")
            return False
        
        logger.info("Loading Raw Data sheet into DataFrame...")
        raw_data_df = pd.DataFrame(wb["Raw Data"].values)
        raw_data_row_count = len(raw_data_df) - 1
        
        raw_data_df.columns = raw_data_df.iloc[0]
        raw_data_df = raw_data_df[1:]
        logger.info(f"Raw Data loaded with {raw_data_row_count} rows and {len(raw_data_df.columns)} columns")
        
        if "Invoice Number" not in raw_data_df.columns and "Receipt Number" not in raw_data_df.columns:
            logger.error("❌ Error: Neither Invoice Number nor Receipt Number column found in Raw Data.")
            return False
        logger.info("✅ Required identifier columns found in Raw Data")
        
        logger.info("Creating unique identifiers for each row...")
        def get_uid(row):
            ttype = str(row.get("Transaction Type", "")).strip().lower()
            if ttype == "invoice" or ttype == "credit memo":
                return row.get("Invoice Number")
            elif ttype == "receipt":
                return row.get("Receipt Number")
            return None
        
        raw_data_df["__UniqueID__"] = raw_data_df.apply(get_uid, axis=1)
        raw_data_df = raw_data_df[raw_data_df["__UniqueID__"].notna()]
        logger.info(f"Created UIDs - {len(raw_data_df)} valid rows with identifiers")
        
        total_sheets = len(wb.sheetnames) - 1
        processed_sheets = 0
        updated_records = 0
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Raw Data":
                continue
                
            logger.info(f"Processing sheet: {sheet_name}")
            processed_sheets += 1
            
            sheet_df = pd.DataFrame(wb[sheet_name].values)
            if len(sheet_df) <= 1:
                logger.info(f"⏭️ Sheet {sheet_name} is empty or contains only headers - skipping")
                continue
                
            sheet_df.columns = sheet_df.iloc[0]
            sheet_df = sheet_df[1:]
            logger.info(f"Sheet {sheet_name} loaded with {len(sheet_df)} rows")
            
            user_cols = []
            if "User Comments" in sheet_df.columns:
                user_cols.append("User Comments")
            if "Status" in sheet_df.columns:
                user_cols.append("Status")
                
            if not user_cols:
                logger.info(f"⏭️ No User Comments or Status columns found in {sheet_name} - skipping")
                continue
            logger.info(f"Found user columns in {sheet_name}: {', '.join(user_cols)}")
                
            sheet_df["__UniqueID__"] = sheet_df.apply(get_uid, axis=1)
            sheet_df = sheet_df[sheet_df["__UniqueID__"].notna()]
            logger.info(f"Created {len(sheet_df)} valid UIDs in {sheet_name}")
            
            sheet_updates = 0
            
            for idx, row in sheet_df.iterrows():
                uid = row["__UniqueID__"]
                if uid:
                    raw_matches = raw_data_df[raw_data_df["__UniqueID__"] == uid].index
                    if len(raw_matches) > 0:
                        for col in user_cols:
                            if pd.notna(row[col]):
                                for match_idx in raw_matches:
                                    raw_data_df.at[match_idx, col] = row[col]
                                    sheet_updates += 1
            
            updated_records += sheet_updates
            logger.info(f"✅ Updated {sheet_updates} values from {sheet_name} to Raw Data")
        
        logger.info(f"Writing {len(raw_data_df)} rows back to Raw Data sheet...")
        raw_sheet = wb["Raw Data"]
        for row in range(2, raw_sheet.max_row + 1):
            for col in range(1, raw_sheet.max_column + 1):
                cell = raw_sheet.cell(row=row, column=col)
                cell.value = None
        logger.info("Cleared existing data in Raw Data sheet")
                
        headers = list(raw_data_df.columns)
        for col_idx, header in enumerate(headers, 1):
            raw_sheet.cell(row=1, column=col_idx, value=header)
            
        for row_idx, (_, row) in enumerate(raw_data_df.iterrows(), 2):
            for col_idx, col_name in enumerate(headers, 1):
                if col_name != "__UniqueID__":  # Don't write the temporary UID column
                    raw_sheet.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        logger.info("Saving updated workbook...")
        wb.save(master_file)
        logger.info(f"✅ Successfully consolidated comments: Processed {processed_sheets} sheets, updated {updated_records} records")
        logger.info(f"Log file saved to: {log_file}")
        return True
        
    except Exception as e:
        logger.error("❌ Error consolidating data")
        logger.exception(e)
        raise  # Let run_app4 handle the popup

#******END OF CONSOLIDATE COMMENTS TO RAW DATA*********



# === GUI Setup ===

# Function for running App 1 (Schedule AR Report)
def run_app1():
    try:
        update_status("Scheduling AR Report - Please wait...")
        getreport_job()
        update_status("AR Report Scheduled.")
        messagebox.showinfo("Success!", "AR Report Scheduled")
    except FileNotFoundError as e:
        update_status("Error: File not found.")
        messagebox.showerror("File Not Found", f"Required file was not found:\n{e}")
    except PermissionError as e:
        update_status("Error: File is open or locked.")
        messagebox.showerror("File Locked", f"File is open in another program or you do not have permission:\n{e}")
    except ValueError as e:
        update_status("Error: Data error.")
        messagebox.showerror("Data Error", f"Data error:\n{e}")
    except Exception as e:
        update_status("An unexpected error occurred.")
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred:\n{e}")
    finally:
        enable_buttons()

# Function for running App 2 (Download AR Report)
def run_app2():
    try:
        update_status("Downloading Report- Please wait...")
        start_Downloadprocess()
        update_status("AR Report Downloaded and split.")
        messagebox.showinfo("Success!", "AR Report Downloaded")
    except FileNotFoundError as e:
        update_status("Error: File not found.")
        messagebox.showerror("File Not Found", f"Required file was not found:\n{e}")
    except PermissionError as e:
        update_status("Error: File is open or locked.")
        messagebox.showerror("File Locked", f"File is open in another program or you do not have permission:\n{e}")
    except ValueError as e:
        update_status("Error: Data error.")
        messagebox.showerror("Data Error", f"Data error:\n{e}")
    except Exception as e:
        update_status("An unexpected error occurred.")
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred:\n{e}")
    finally:
        enable_buttons()

# Function for running App 3 (Update Master File)
def run_app3():
    try:
        update_status("Updating Master File - Please wait...")
        run_Master_file_script()
        update_status("Master file updated.")
        messagebox.showinfo("Success!", "Master file updated")
    except FileNotFoundError as e:
        update_status("Error: File not found.")
        messagebox.showerror("File Not Found", f"Required file was not found:\n{e}")
    except PermissionError as e:
        update_status("Error: File is open or locked.")
        messagebox.showerror("File Locked", f"File is open in another program or you do not have permission:\n{e}")
    except ValueError as e:
        update_status("Error: Data error.")
        messagebox.showerror("Data Error", f"Data error:\n{e}")
    except Exception as e:
        update_status("An unexpected error occurred.")
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred:\n{e}")
    finally:
        enable_buttons()

# Function for running App 4 (Consolidate Comments to Raw Data)
def run_app4():
    try:
        update_status("Consolidating Comments to Raw Data - Please wait...")
        consolidate_comments_to_raw_data()
        update_status("Comments consolidated to Raw Data.")
        messagebox.showinfo("Success!", "Comments consolidated to Raw Data")
    except FileNotFoundError as e:
        update_status("Error: File not found.")
        messagebox.showerror("File Not Found", f"Required file was not found:\n{e}")
    except PermissionError as e:
        update_status("Error: File is open or locked.")
        messagebox.showerror("File Locked", f"File is open in another program or you do not have permission:\n{e}")
    except ValueError as e:
        update_status("Error: Data error.")
        messagebox.showerror("Data Error", f"Data error:\n{e}")
    except Exception as e:
        update_status("An unexpected error occurred.")
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred:\n{e}")
    finally:
        enable_buttons()

# Update status text in the status label
def update_status(status_message):
    status_label.config(text=status_message)

# Disable all buttons while an app is running
def disable_buttons():
    app1_button.config(state=tk.DISABLED)
    app2_button.config(state=tk.DISABLED)
    app3_button.config(state=tk.DISABLED)
    #app4_button.config(state=tk.DISABLED)

# Enable all buttons after app finishes
def enable_buttons():
    app1_button.config(state=tk.NORMAL)
    app2_button.config(state=tk.NORMAL)
    app3_button.config(state=tk.NORMAL)
    #app4_button.config(state=tk.NORMAL)

# Run the selected app in a separate thread
def run_in_thread(app_function):
    disable_buttons()
    thread = Thread(target=app_function)
    thread.daemon = True  # Ensure thread exits when the main program exits
    thread.start()

# Close the window (Terminate the app)
def on_close():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.quit()

# === GUI Logic ===

# Create the main window
root = tk.Tk()
root.title("AR Aging App V1.4.1 - 14.Jul.25")
root.iconbitmap(resource_path("myicon.ico"))  # Set the icon for the window
root.geometry("400x220")

# Create status label
status_label = tk.Label(root, text="Welcome! Please select a stage to run.", font=("Arial", 12, "bold"), anchor="w", padx=10, pady=10)
status_label.pack(fill=tk.X)

# Create buttons for running each app
app1_button = tk.Button(root, text="1. Schedule AR Report", font=("Arial", 12), width=35, command=lambda: run_in_thread(run_app1))
app1_button.pack(pady=5)

app2_button = tk.Button(root, text="2. Download AR Report", font=("Arial", 12), width=35, command=lambda: run_in_thread(run_app2))
app2_button.pack(pady=5)

app3_button = tk.Button(root, text="3. Update Master file", font=("Arial", 12), width=35, command=lambda: run_in_thread(run_app3))
app3_button.pack(pady=5)

#app4_button = tk.Button(root, text="Consolidate Comments to Raw Data", font=("Arial", 12), width=35, command=lambda: run_in_thread(run_app4))
#app4_button.pack(pady=5)

# Close event
root.protocol("WM_DELETE_WINDOW", on_close)

# Start the GUI main loop
root.mainloop()
